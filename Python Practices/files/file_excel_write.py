import openpyxl

wb=openpyxl.Workbook()
sheet=wb.active

c1=sheet.cell(row=1,column=1)
c1.value='Hello'

c2=sheet.cell(row=1,column=2)
c2.value='world'

c3=sheet['A2']
c3.value='welcome'

c4=sheet['B2']
c4.value='everyone'

wb.save('sample.xlsx')
