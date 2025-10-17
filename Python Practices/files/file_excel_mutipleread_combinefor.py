import openpyxl

path='student.xlsx'
wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active

row=sheet_obj.max_row
column=sheet_obj.max_column

for i in range(1,row+1):
    for j in range(1,column+1):
      cell_obj=sheet_obj.cell(row=i,column=j)
      print(cell_obj.value,end='')

    
